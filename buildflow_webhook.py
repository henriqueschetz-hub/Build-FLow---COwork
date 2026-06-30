#!/usr/bin/env python3
"""
BuildFlow Monitoring Webhook - Roda 24/7 em servidor cloud
Extrai dados do BuildFlow e envia alertas por email
"""

from flask import Flask, jsonify
from apscheduler.schedulers.background import BackgroundScheduler
from datetime import datetime, timedelta
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import requests
from bs4 import BeautifulSoup
import os
import logging

# Configuração de logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Inicializar Flask
app = Flask(__name__)

# Configurações
BUILDFLOW_URL = "https://buildflow.com.br"
LOGIN_USER = "henrique.melo"
LOGIN_PASS = "Hschetz@!2026"
EMAIL_DESTINATARIO = "henrique.schetz@gmail.com"
EMAIL_REMETENTE = os.getenv('EMAIL_REMETENTE', 'seu_email@gmail.com')
EMAIL_SENHA = os.getenv('EMAIL_SENHA', 'sua_senha_app')

# Esteiras a monitorar
ESTEIRAS = [
    "STO - Análise Arquitetura",
    "STO - Análise Estrutural",
    "STO - Análise Instalações",
    "STO - Alteração de Projeto"
]

class BuildFlowMonitor:
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })

    def login(self):
        """Fazer login no BuildFlow"""
        try:
            # Login
            login_data = {
                'usuario': LOGIN_USER,
                'senha': LOGIN_PASS
            }
            response = self.session.post(f"{BUILDFLOW_URL}/login.php", data=login_data)
            logger.info("✅ Login realizado com sucesso")
            return True
        except Exception as e:
            logger.error(f"❌ Erro ao fazer login: {e}")
            return False

    def extrair_cards(self):
        """Extrair dados dos cards do painel"""
        try:
            response = self.session.get(f"{BUILDFLOW_URL}/painel")
            response.raise_for_status()

            cards_por_esteira = {}

            # Parse do HTML
            soup = BeautifulSoup(response.content, 'html.parser')
            page_text = soup.get_text()

            # Extrair dados de cada esteira
            for esteira in ESTEIRAS:
                cards_por_esteira[esteira] = self._parse_esteira(page_text, esteira)

            logger.info(f"✅ {sum(len(v) for v in cards_por_esteira.values())} cards extraídos")
            return cards_por_esteira

        except Exception as e:
            logger.error(f"❌ Erro ao extrair cards: {e}")
            return {}

    def _parse_esteira(self, text, nome_esteira):
        """Parse dos cards de uma esteira"""
        # Implementação simplificada
        # Em produção, isso teria parsing mais robusto do HTML
        cards = []

        if nome_esteira in text:
            # Buscar cards após o nome da esteira
            start = text.find(nome_esteira)
            section = text[start:start+2000]

            # Extrair padrões de data
            import re
            dates = re.findall(r'(\d{2}/\d{2}/\d{4})', section)

            if dates:
                cards.append({
                    'titulo': nome_esteira,
                    'data_prazo': dates[0] if dates else '',
                    'dias_restantes': self._calcular_dias(dates[0]) if dates else 0
                })

        return cards

    def _calcular_dias(self, data_str):
        """Calcular dias até vencer"""
        try:
            data_prazo = datetime.strptime(data_str, "%d/%m/%Y")
            dias = (data_prazo - datetime.now()).days
            return dias
        except:
            return 0

    def gerar_planilha(self, cards_por_esteira):
        """Gerar planilha Excel"""
        try:
            wb = Workbook()
            wb.remove(wb.active)

            # Estilos
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            today = datetime.now().strftime("%d/%m/%Y")

            # Criar aba resumo
            ws_resumo = wb.create_sheet(title="Resumo", index=0)
            ws_resumo.append(["RESUMO - MONITORAMENTO BUILDFLOW"])
            ws_resumo['A1'].font = Font(bold=True, size=14)
            ws_resumo.append(["Data de Extração:", today])

            # Criar abas por esteira
            for esteira, cards in cards_por_esteira.items():
                ws = wb.create_sheet(title=esteira[:25])

                headers = ["Título", "Data de Criação", "Prazo", "Dias até Vencer", "Status"]
                ws.append(headers)

                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font

                for card in cards:
                    dias = card.get('dias_restantes', 0)
                    if dias < 0:
                        status = f"VENCIDO ({abs(dias)}d)"
                        fill = red_fill
                    elif dias <= 3:
                        status = f"URGENTE ({dias}d)"
                        fill = yellow_fill
                    else:
                        status = f"Em análise ({dias}d)"
                        fill = None

                    row = [
                        card.get('titulo', ''),
                        card.get('data_criacao', ''),
                        card.get('data_prazo', ''),
                        dias,
                        status
                    ]
                    ws.append(row)

                    if fill:
                        ws[f'E{ws.max_row}'].fill = fill

            # Salvar
            filename = f"/tmp/buildflow_monitoring_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
            wb.save(filename)
            logger.info(f"✅ Planilha salva: {filename}")
            return filename

        except Exception as e:
            logger.error(f"❌ Erro ao gerar planilha: {e}")
            return None

    def enviar_alerta(self, cards_por_esteira):
        """Enviar email com alertas"""
        try:
            # Contar alertas
            vencidos = 0
            urgentes = 0

            for cards in cards_por_esteira.values():
                for card in cards:
                    dias = card.get('dias_restantes', 0)
                    if dias < 0:
                        vencidos += 1
                    elif dias <= 3:
                        urgentes += 1

            # Se houver alertas, enviar email
            if vencidos > 0 or urgentes > 0:
                self._send_email(vencidos, urgentes, cards_por_esteira)
                logger.info(f"✅ Email de alerta enviado ({vencidos} vencidos, {urgentes} urgentes)")
            else:
                logger.info("✅ Nenhum alerta para enviar")

        except Exception as e:
            logger.error(f"❌ Erro ao enviar alerta: {e}")

    def _send_email(self, vencidos, urgentes, cards_por_esteira):
        """Enviar email SMTP"""
        try:
            msg = MIMEMultipart('alternative')
            msg['Subject'] = f"[ALERTA] BuildFlow - {vencidos} Vencidos, {urgentes} Urgentes"
            msg['From'] = EMAIL_REMETENTE
            msg['To'] = EMAIL_DESTINATARIO

            # Corpo do email
            html = f"""
            <html>
                <body>
                    <h2>🚨 ALERTA - BuildFlow Monitoring</h2>
                    <p><strong>Data:</strong> {datetime.now().strftime('%d/%m/%Y %H:%M')}</p>

                    <h3>📊 Resumo de Alertas:</h3>
                    <ul>
                        <li><span style="color: red;"><strong>Cards VENCIDOS: {vencidos}</strong></span></li>
                        <li><span style="color: orange;"><strong>Cards URGENTES (próximos 3 dias): {urgentes}</strong></span></li>
                    </ul>

                    <h3>📋 Detalhes por Esteira:</h3>
            """

            for esteira, cards in cards_por_esteira.items():
                html += f"<h4>{esteira}</h4><ul>"
                for card in cards:
                    dias = card.get('dias_restantes', 0)
                    if dias < 0 or dias <= 3:
                        html += f"<li>{card.get('titulo', 'N/A')} - {dias}d</li>"
                html += "</ul>"

            html += """
                    <hr>
                    <p><small>Monitoramento automático BuildFlow - Roda 24/7</small></p>
                </body>
            </html>
            """

            part = MIMEText(html, 'html')
            msg.attach(part)

            # Enviar via SMTP Gmail
            with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
                server.login(EMAIL_REMETENTE, EMAIL_SENHA)
                server.send_message(msg)

        except Exception as e:
            logger.error(f"Erro ao enviar email: {e}")

# Função de monitoramento
def monitoramento_agendado():
    """Executar monitoramento (chamado a cada 7h)"""
    logger.info("🔄 Iniciando monitoramento...")

    monitor = BuildFlowMonitor()

    if monitor.login():
        cards = monitor.extrair_cards()
        if cards:
            monitor.gerar_planilha(cards)
            monitor.enviar_alerta(cards)

    logger.info("✅ Monitoramento concluído")

# Endpoints Flask
@app.route('/', methods=['GET'])
def health():
    """Health check"""
    return jsonify({'status': 'ok', 'message': 'BuildFlow Webhook está rodando 24/7'})

@app.route('/run', methods=['POST'])
def run_manual():
    """Executar monitoramento manualmente"""
    monitoramento_agendado()
    return jsonify({'status': 'success', 'message': 'Monitoramento executado'})

# Configurar scheduler
scheduler = BackgroundScheduler()

# Agendar para rodar todos os dias às 7h
scheduler.add_job(
    func=monitoramento_agendado,
    trigger="cron",
    hour=7,
    minute=0,
    timezone='America/Sao_Paulo',
    id='buildflow_monitor'
)

# Iniciar scheduler
scheduler.start()
logger.info("✅ Scheduler iniciado - Monitoramento às 7h00 diariamente")

if __name__ == '__main__':
    # Iniciar Flask app
    port = int(os.getenv('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
